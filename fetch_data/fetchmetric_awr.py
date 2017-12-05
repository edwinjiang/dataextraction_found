#!/usr/bin/python
#encoding:utf8

from __future__ import print_function

from bs4 import BeautifulSoup
import os
import re
import argparse
import json
from openpyxl import load_workbook
import glob


DATADIR = ""
awrfile = ""
awrdir = ""

def extrace_from_file(awrfile,line_no):
    with open(awrfile) as f:
        soup = BeautifulSoup(f, 'lxml')

        # Oracle Release Version
        dbrelease = []

        for rawdata in soup.find_all('table'):
            if re.compile('Database Instances Included In Report').search(rawdata['summary']):
                trdatas = rawdata.find_all(name='tr')
                for trdata in trdatas:
                    tddatas = BeautifulSoup(str(trdata), "html.parser")
                    tddata = tddatas.find_all(name='td')
                    #for tddata in tddatas.find_all(name='td'):
        dbrelease.append(tddata[6].string)
        # print dbname[0]["DB Name"]
        # print begin[0]["Begin SnapShot"]
        # print endy[0]["End Snapshot"]
        f.close()

        # DB Name
        dbname = []
        begin = []
        endy = []

        for rawdata in soup.find_all('table'):
            if re.compile('Database Summary').search(rawdata['summary']):
                trdatas = rawdata.find_all(name='tr')
                for trdata in trdatas:
                    tddatas = BeautifulSoup(str(trdata), "html.parser")
                    for tddata in tddatas.find_all(name='td'):
                        if tddata.get('headers') != None:
                            if re.compile('Name').search(tddata.get('headers')[1]):
                                dbname.append({"DB Name": tddata.string})
                            if re.compile('Begin').search(tddata.get('headers')[1]):
                                begin.append({"Begin SnapShot": tddata.string})
                            if re.compile('End').search(tddata.get('headers')[1]):
                                endy.append({"End Snapshot": tddata.string})
        # print dbname[0]["DB Name"]
        # print begin[0]["Begin SnapShot"]
        # print endy[0]["End Snapshot"]
        f.close()

    # DB_CACHE_SIZE
    with open(awrfile) as f:
        soup = BeautifulSoup(f, 'lxml')
        dbcache = []
        dbcachesize = 0
        try:
            for rawdata in soup.find_all('table'):
                if re.compile('This table displays values for init.ora parameters').search(rawdata['summary']):
                    trdatas = rawdata.find_all(name='tr')
                    for trdata in trdatas:
                        tddatas = BeautifulSoup(str(trdata), "html.parser")
                        tddata = tddatas.find_all(name='td')
                        try:
                            if (re.compile('db_cache_size').match(tddata[0].string)):
                                dbcache.append({'db cache size': tddata[2].string})
                        except Exception, e:
                            pass
            dbcachesize =  float(dbcache[0]['db cache size'])/1024/1024/1024
            f.close()
        except Exception, e:
            pass

    # IOPS Metircs
    with open(awrfile) as f:
        soup = BeautifulSoup(f, 'lxml')

        io_data = []
        result = {}
        total_iops = 0
        total_throughput = 0

        pattern_io = re.compile('physical')
        pattern_io_iops_metric = re.compile('total IO requests')
        # Throughput metric
        pattern_io_throughput_metric = re.compile('total bytes')

        for rawdata in soup.find_all('table'):
            if re.compile('System Statistics').search(rawdata['summary']):
                trdatas = rawdata.find_all(name='tr')
                for trdata in trdatas:
                    tddatas = BeautifulSoup(str(trdata), "html.parser")
                    tddata = tddatas.find_all(name='td')
                    try:
                        if (pattern_io.match(tddata[0].string)):
                            io_data.append({'name': tddata[0].string, 'total': float(tddata[1].string.replace(',', '')),
                                            'per second': float(tddata[2].string.replace(',', '')),
                                            'per trans': float(tddata[3].string.replace(',', ''))})
                    except Exception, e:
                        pass

        for i in range(len(io_data)):
            if pattern_io_iops_metric.search(io_data[i]["name"]):
                total_iops += io_data[i]["per second"]
                result[io_data[i]["name"]] = io_data[i]["per second"]
            if pattern_io_throughput_metric.search(io_data[i]["name"]):
                total_throughput += io_data[i]["per second"]
                result[io_data[i]["name"]] = io_data[i]["per second"]
        result["Total IOPS"] = total_iops
        result["Total ThroughPut"] = total_throughput
        json_io = json.dumps(result)
        # print result
        # pprint.pprint(json_io)
        # print result["physical write total IO requests"]
        f.close()

        # CPU Metrics
        with open(awrfile) as f:
            soup = BeautifulSoup(f, 'lxml')
            # CPU metric
            pattern_cpu_Busy = re.compile('%Busy')
            pattern_cpu_Usr = re.compile('%Usr')
            pattern_cpu_Sys = re.compile('%Sys')
            cpu_data = []
            instance_info = ""
            cpu_data_dic = {}

            if dbrelease[0] == "12.2.0.1.0":
                try:
                    for rawdata in soup.find_all('table'):
                        if re.compile('OS Statistics By Instance').search(rawdata['summary']):
                            trdatas = rawdata.find_all(name='tr')
                            for trdata in trdatas:
                                tddatas = BeautifulSoup(str(trdata), "html.parser")
                                for tddata in tddatas.find_all(name='td'):
                                    # if tddata.get('scope') != None:
                                    #     if re.compile('row').search(tddata.get('scope')):
                                    #         instance = tddata.string
                                    #         cpu_data.append({'Instance#': instance})
                                    if tddata.get('headers') != None:
                                        if pattern_cpu_Busy.search(tddata.get('headers')[1]):
                                            cpu_data.append({'%Busy': tddata.string})
                                        if pattern_cpu_Usr.search(tddata.get('headers')[1]):
                                            cpu_data.append({'%Usr': tddata.string})
                                        if pattern_cpu_Sys.search(tddata.get('headers')[1]):
                                            cpu_data.append({'%Sys': tddata.string})
                    f.close()
                except Exception, e:
                    pass
                #print (cpu_data)

            if dbrelease[0] == "12.1.0.2.0":
                cpu_data1 = []
                try:
                    for rawdata in soup.find_all('table'):
                        if re.compile('OS Statistics By Instance').search(rawdata['summary']):
                            trdatas = rawdata.find_all(name='tr')
                            for trdata in trdatas:
                                tddatas = BeautifulSoup(str(trdata), "html.parser")
                                for tddata in tddatas.find_all(name='td'):
                                    cpu_data1.append(tddata)
                            tddatas1 = BeautifulSoup(str(cpu_data1[6]), "html.parser")
                            tddata1 = tddatas1.find_all(name='td')
                            cpu_data.append({'%Busy': tddata1[0].string})
                            tddatas1 = BeautifulSoup(str(cpu_data1[7]), "html.parser")
                            tddata1 = tddatas1.find_all(name='td')
                            cpu_data.append({'%Usr': tddata1[0].string})
                            tddatas1 = BeautifulSoup(str(cpu_data1[8]), "html.parser")
                            tddata1 = tddatas1.find_all(name='td')
                            cpu_data.append({'%Sys': tddata1[0].string})

                            tddatas1 = BeautifulSoup(str(cpu_data1[25]), "html.parser")
                            tddata1 = tddatas1.find_all(name='td')
                            cpu_data.append({'%Busy': tddata1[0].string})
                            tddatas1 = BeautifulSoup(str(cpu_data1[26]), "html.parser")
                            tddata1 = tddatas1.find_all(name='td')
                            cpu_data.append({'%Usr': tddata1[0].string})
                            tddatas1 = BeautifulSoup(str(cpu_data1[27]), "html.parser")
                            tddata1 = tddatas1.find_all(name='td')
                            cpu_data.append({'%Sys': tddata1[0].string})
                    f.close()
                except Exception, e:
                    pass
                #print (cpu_data)

        # Foregroud Wait Time Metrics
        with open(awrfile) as f:
            soup = BeautifulSoup(f, 'lxml')

            # IO latency calculation
            db_file_parallel_read = re.compile('db file parallel read')
            db_file_sequential_read = re.compile('db file sequential read')
            control_file_sequential_read = re.compile('control file sequential read')

            db_file_parallel_read_wait = []
            db_file_parallel_read_wait_time = []
            db_file_parallel_read_wait_dic = {}

            db_file_sequential_read_wait = []
            db_file_sequential_read_wait_time = []
            db_file_sequential_read_wait_dic = {}

            control_file_sequential_read_wait = []
            control_file_sequential_read_wait_time = []
            control_file_sequential_read_dic = {}


            for rawdata in soup.find_all('table'):
                if re.compile('Top Timed Foreground Events. . Foreground Activity ').search(rawdata['summary']):
                    trdatas = rawdata.find_all(name='tr')
                    for trdata in trdatas:
                        tddatas = BeautifulSoup(str(trdata), "html.parser")
                        tddata = tddatas.find_all(name='td')
                        try:
                            if (db_file_parallel_read.match(tddata[2].string)):
                                db_file_parallel_read_wait.append(tddata[3].string)
                                db_file_parallel_read_wait_time.append(tddata[6].string)

                            if (db_file_sequential_read.match(tddata[2].string)):
                                db_file_sequential_read_wait.append(tddata[3].string)
                                db_file_sequential_read_wait_time.append(tddata[6].string)

                            if (control_file_sequential_read.match(tddata[2].string)):
                                control_file_sequential_read_wait.append(tddata[3].string)
                                control_file_sequential_read_wait_time.append(tddata[6].string)

                        except Exception, e:
                            pass
            try:
                # print foregroud_waittime
                if len(db_file_sequential_read_wait) == 0:
                    db_file_sequential_read_wait.append(0)
                if len(db_file_sequential_read_wait_time) == 0:
                    db_file_sequential_read_wait_time.append(0)
                db_file_sequential_read_wait_dic["db file sequential read Waits"]=db_file_sequential_read_wait
                db_file_sequential_read_wait_dic["db file sequential read Avg Wait time"] = db_file_sequential_read_wait_time


                if len(db_file_parallel_read_wait) == 0:
                    db_file_parallel_read_wait.append(0)
                if len(db_file_parallel_read_wait_time) == 0:
                    db_file_parallel_read_wait_time.append(0)
                db_file_parallel_read_wait_dic["db file parallel read Waits"] = db_file_parallel_read_wait
                db_file_parallel_read_wait_dic["db file parallel read Avg Wait time"] = db_file_parallel_read_wait_time

                if len(control_file_sequential_read_wait) == 0:
                    control_file_sequential_read_wait.append(0)
                if len(control_file_sequential_read_wait_time) == 0:
                    control_file_sequential_read_wait_time.append(0)
                control_file_sequential_read_dic["control file sequential read write Waits"] = control_file_sequential_read_wait
                control_file_sequential_read_dic["control file sequential read Avg Wait time"] = control_file_sequential_read_wait_time

                # print(db_file_sequential_read_wait_dic["db file sequential read Waits"][0])
                # print(db_file_sequential_read_wait_dic["db file sequential read Avg Wait time"][0])
                #
                # print(db_file_parallel_read_wait_dic["db file parallel read Waits"][0])
                # print(db_file_parallel_read_wait_dic["db file parallel read Avg Wait time"][0])
                #
                # print(control_file_sequential_read_dic["control file sequential read write Waits"][0])
                # print(control_file_sequential_read_dic["control file sequential read Avg Wait time"][0])
            except Exception, e:
                pass
            f.close()

            # Backgroud Wait Time Metrics
            with open(awrfile) as f:
                soup = BeautifulSoup(f, 'lxml')

                # IO latency calculation
                log_file_parallel_write = re.compile('log file parallel write')
                control_file_parallel_write = re.compile('control file parallel write')
                db_file_paralle_write = re.compile('db file parallel write')

                log_file_parallel_write_wait = []
                log_file_parallel_write_wait_time = []
                log_file_parallel_write_wait_dic = {}

                control_file_parallel_write_wait = []
                control_file_parallel_write_wait_time = []
                control_file_parallel_write_wait_dic = {}

                db_file_paralle_write_wait = []
                db_file_paralle_write_wait_time = []
                db_file_paralle_write_wait_dic = {}

                for rawdata in soup.find_all('table'):
                    if re.compile('Top Timed Background Events').search(rawdata['summary']):
                        trdatas = rawdata.find_all(name='tr')
                        for trdata in trdatas:
                            tddatas = BeautifulSoup(str(trdata), "html.parser")
                            tddata = tddatas.find_all(name='td')
                            try:
                                if (log_file_parallel_write.match(tddata[2].string)):
                                    log_file_parallel_write_wait.append(tddata[3].string)
                                    log_file_parallel_write_wait_time.append(tddata[6].string)

                                if (control_file_parallel_write.match(tddata[2].string)):
                                    control_file_parallel_write_wait.append(tddata[3].string)
                                    control_file_parallel_write_wait_time.append(tddata[6].string)

                                if (db_file_paralle_write.match(tddata[2].string)):
                                    db_file_paralle_write_wait.append(tddata[3].string)
                                    db_file_paralle_write_wait_time.append(tddata[6].string)

                            except Exception, e:
                                pass
                # print backgroud_waittime
                if len(log_file_parallel_write_wait) == 0:
                    log_file_parallel_write_wait.append(0)
                if len(log_file_parallel_write_wait_time) == 0:
                    log_file_parallel_write_wait_time.append(0)
                log_file_parallel_write_wait_dic["log file parallel write Waits"] = log_file_parallel_write_wait
                log_file_parallel_write_wait_dic["log file parallel write Avg Wait time"] = log_file_parallel_write_wait_time

                if len(control_file_parallel_write_wait) == 0:
                    control_file_parallel_write_wait.append(0)
                if len(control_file_parallel_write_wait_time) == 0:
                    control_file_parallel_write_wait_time.append(0)
                control_file_parallel_write_wait_dic["control file parallel write write Waits"] = control_file_parallel_write_wait
                control_file_parallel_write_wait_dic["control file parallel write Avg Wait time"] = control_file_parallel_write_wait_time

                if len(db_file_paralle_write_wait) == 0:
                    db_file_paralle_write_wait.append(0)
                if len(db_file_paralle_write_wait_time) == 0:
                    db_file_paralle_write_wait_time.append(0)
                db_file_paralle_write_wait_dic["db file parallel write write Waits"] = db_file_paralle_write_wait
                db_file_paralle_write_wait_dic["db file parallel write Avg Wait time"] = db_file_paralle_write_wait_time

                f.close()
                print (dbname[0]["DB Name"]+" AWR Report parsing is finished")

    try:
        wb = load_workbook("awr_metrics.xlsx")
        ws = wb.get_sheet_by_name("awr_metric")
        #print line_no
        ws.cell(row=line_no, column=3).value = dbname[0]["DB Name"]
        ws.cell(row=line_no, column=4).value = begin[0]["Begin SnapShot"] + "--" + endy[0]["End Snapshot"]
        ws.cell(row=line_no, column=6).value = dbcachesize
        if len(cpu_data) > 3:
            for i in range(len(cpu_data) - 1):
                for key, value in cpu_data[i].items():
                    ws.cell(row=line_no, column=7).value = cpu_data[0]["%Busy"]
                    ws.cell(row=line_no, column=8).value = cpu_data[3]["%Busy"]
        if len(cpu_data) <= 3:
            print (dbname[0]["DB Name"]+" AWR report Loss One Node CPU data")
            ws.cell(row=line_no, column=7).value = 0
            ws.cell(row=line_no, column=8).value = 0

        ws.cell(row=line_no, column=9).value = float(result["physical read total IO requests"])
        ws.cell(row=line_no, column=10).value = float(result["physical write total IO requests"])
        ws.cell(row=line_no, column=11).value = float(result["physical read total IO requests"]) + float(
            result["physical write total IO requests"])
        ws.cell(row=line_no, column=12).value = db_file_sequential_read_wait_dic["db file sequential read Waits"][0]
        ws.cell(row=line_no, column=13).value = db_file_sequential_read_wait_dic["db file sequential read Avg Wait time"][0]
        ws.cell(row=line_no, column=14).value = db_file_parallel_read_wait_dic["db file parallel read Waits"][0]
        ws.cell(row=line_no, column=15).value = db_file_parallel_read_wait_dic["db file parallel read Avg Wait time"][0]
        ws.cell(row=line_no, column=16).value = log_file_parallel_write_wait_dic["log file parallel write Waits"][0]
        ws.cell(row=line_no, column=17).value = log_file_parallel_write_wait_dic["log file parallel write Avg Wait time"][0]
        ws.cell(row=line_no, column=18).value = db_file_paralle_write_wait_dic["db file parallel write write Waits"][0]
        ws.cell(row=line_no, column=19).value = db_file_paralle_write_wait_dic["db file parallel write Avg Wait time"][0]
        wb.save("awr_metrics.xlsx")
    except Exception, e:
        pass

def extrace_from_directory(awrdir):

    files = glob.iglob(awrdir+'\*.html')

    # The actual data in Excel Template at the third row
    line_no = 3
    for item in files:
        extrace_from_file(item,line_no)
        line_no +=1

def clear_data():
    try:
        wb = load_workbook("awr_metrics.xlsx")
        ws = wb.get_sheet_by_name("awr_metric")
        # clear excel data
        maxrow = ws.max_row
        for i in range(3,maxrow+1):
            ws.cell(row=i, column=3).value = ''
            ws.cell(row=i, column=4).value = ''
            ws.cell(row=i, column=5).value = ''
            ws.cell(row=i, column=6).value = ''
            ws.cell(row=i, column=7).value = ''
            ws.cell(row=i, column=8).value = ''
            ws.cell(row=i, column=9).value = ''
            ws.cell(row=i, column=10).value = ''
            ws.cell(row=i, column=11).value = ''
            ws.cell(row=i, column=12).value = ''
            ws.cell(row=i, column=13).value = ''
            ws.cell(row=i, column=14).value = ''
            ws.cell(row=i, column=15).value = ''
            ws.cell(row=i, column=16).value = ''
            ws.cell(row=i, column=17).value = ''
            ws.cell(row=i, column=18).value = ''
            ws.cell(row=i, column=19).value = ''
        wb.save("awr_metrics.xlsx")
    except Exception, e:
        print (e)

def extract_data(**kwargs):

    if kwargs["file"] is not None:
        awrfile = os.path.join(DATADIR, kwargs["file"])
        extrace_from_file(awrfile,3)
    if  kwargs["directory"] is not None:
        awrdir = kwargs["directory"]
        extrace_from_directory(awrdir)


def _argparser():
    parser = argparse.ArgumentParser(description='AWR Data Analyze')
    parser.add_argument('--f','--file', action='store', dest='file', help='AWR File Name')
    parser.add_argument('--d', '--directory', action='store', dest='directory', help='AWR File Directory')

    return parser.parse_args()

if __name__ == "__main__":
    parser = _argparser()
    if parser.file is None and parser.directory is None:
        print ("usage: fetchmetric_awr.py [-h] [--f FILE] [--d DIRECTORY]")
    conn_args = dict(file=parser.file,directory=parser.directory)
    clear_data()
    extract_data(**conn_args)

