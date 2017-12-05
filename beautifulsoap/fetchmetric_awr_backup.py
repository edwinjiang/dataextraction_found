#!/usr/bin/python
#encoding:utf8

from bs4 import BeautifulSoup
import os
#import pprint
import re
import argparse
import json
import csv
from openpyxl import load_workbook
import glob

DATADIR = ""
DATAFILE = ""
awrfile = ""
awrdir = ""

def extrace_from_file(awrfile,line_no):
    with open(awrfile) as f:
        soup = BeautifulSoup(f, 'lxml')

        # DB Name
        dbname = []
        begin = []
        endy = []

        for rawdata in soup.find_all('table'):
            if re.compile('Database Summary').search(rawdata['summary']):
                trdatas = rawdata.find_all(name='tr')
                for trdata in trdatas:
                    tddatas = BeautifulSoup(str(trdata), "html.parser")
                    for tddata in tddatas.find_all(name='td'):
                        if tddata.get('headers') != None:
                            if re.compile('Name').search(tddata.get('headers')[1]):
                                dbname.append({"DB Name": tddata.string})
                            if re.compile('Begin').search(tddata.get('headers')[1]):
                                begin.append({"Begin SnapShot": tddata.string})
                            if re.compile('End').search(tddata.get('headers')[1]):
                                endy.append({"End Snapshot": tddata.string})
        # print dbname[0]["DB Name"]
        # print begin[0]["Begin SnapShot"]
        # print endy[0]["End Snapshot"]
        f.close()

    # DB_CACHE_SIZE
    with open(awrfile) as f:
        soup = BeautifulSoup(f, 'lxml')
        dbcache = []
        dbcachesize = 0
        try:
            for rawdata in soup.find_all('table'):
                if re.compile('This table displays values for init.ora parameters').search(rawdata['summary']):
                    trdatas = rawdata.find_all(name='tr')
                    for trdata in trdatas:
                        tddatas = BeautifulSoup(str(trdata), "html.parser")
                        tddata = tddatas.find_all(name='td')
                        try:
                            if (re.compile('db_cache_size').match(tddata[0].string)):
                                dbcache.append({'db cache size': tddata[2].string})
                        except Exception, e:
                            pass
            dbcachesize =  float(dbcache[0]['db cache size'])/1024/1024/1024
            f.close()
        except Exception, e:
            pass

    # IOPS Metircs
    with open(awrfile) as f:
        soup = BeautifulSoup(f, 'lxml')

        io_data = []
        result = {}
        total_iops = 0
        total_throughput = 0

        pattern_io = re.compile('physical')
        pattern_io_iops_metric = re.compile('total IO requests')
        # Throughput metric
        pattern_io_throughput_metric = re.compile('total bytes')

        for rawdata in soup.find_all('table'):
            if re.compile('System Statistics').search(rawdata['summary']):
                trdatas = rawdata.find_all(name='tr')
                for trdata in trdatas:
                    tddatas = BeautifulSoup(str(trdata), "html.parser")
                    tddata = tddatas.find_all(name='td')
                    try:
                        if (pattern_io.match(tddata[0].string)):
                            io_data.append({'name': tddata[0].string, 'total': float(tddata[1].string.replace(',', '')),
                                            'per second': float(tddata[2].string.replace(',', '')),
                                            'per trans': float(tddata[3].string.replace(',', ''))})
                    except Exception, e:
                        pass
        # print data
        for i in range(len(io_data)):
            if pattern_io_iops_metric.search(io_data[i]["name"]):
                total_iops += io_data[i]["per second"]
                result[io_data[i]["name"]] = io_data[i]["per second"]
            if pattern_io_throughput_metric.search(io_data[i]["name"]):
                total_throughput += io_data[i]["per second"]
                result[io_data[i]["name"]] = io_data[i]["per second"]
        result["Total IOPS"] = total_iops
        result["Total ThroughPut"] = total_throughput
        json_io = json.dumps(result)
        # print result
        # pprint.pprint(json_io)
        # print result["physical write total IO requests"]
        f.close()

        # CPU Metrics
        with open(awrfile) as f:
            soup = BeautifulSoup(f, 'lxml')
            # CPU metric
            pattern_cpu_Busy = re.compile('%Busy')
            pattern_cpu_Usr = re.compile('%Usr')
            pattern_cpu_Sys = re.compile('%Sys')
            cpu_data = []
            instance_info = ""
            try:
                for rawdata in soup.find_all('table'):
                    if re.compile('OS Statistics By Instance').search(rawdata['summary']):
                        trdatas = rawdata.find_all(name='tr')
                        for trdata in trdatas:
                            tddatas = BeautifulSoup(str(trdata), "html.parser")
                            for tddata in tddatas.find_all(name='td'):
                                # print tddata
                                if tddata.get('scope') != None:
                                    if re.compile('row').search(tddata.get('scope')):
                                        instance = tddata.string
                                        cpu_data.append({'Instance#': instance})
                                # if tddata.get('headers') != None:
                                #     if pattern_cpu_Busy.search(tddata.get('headers')[1]):
                                #         cpu_data["Instance#"]=instance
                                #         cpu_data["%Busy"]=tddata.string
                                #     if pattern_cpu_Usr.search(tddata.get('headers')[1]):
                                #         cpu_data["Instance#"]=instance
                                #         cpu_data["%Usr"]=tddata.string
                                #     if pattern_cpu_Sys.search(tddata.get('headers')[1]):
                                #         cpu_data["Instance#"]=instance
                                #         cpu_data["%Sys"]=tddata.string

                                if tddata.get('headers') != None:
                                    if pattern_cpu_Busy.search(tddata.get('headers')[1]):
                                        cpu_data.append({'%Busy': tddata.string})
                                    if pattern_cpu_Usr.search(tddata.get('headers')[1]):
                                        cpu_data.append({'%Usr': tddata.string})
                                    if pattern_cpu_Sys.search(tddata.get('headers')[1]):
                                        cpu_data.append({'%Sys': tddata.string})
                #print cpu_data
                # cpudata =[]
                # for i in range(len(cpu_data)):
                #     for key,value in cpu_data[i].items():
                #         print key,value
                f.close()
            except Exception, e:
                pass

        # Foregroud Wait Time Metrics
        with open(awrfile) as f:
            soup = BeautifulSoup(f, 'lxml')

            # IO latency calculation
            db_file_parallel_read = re.compile('db file parallel read')
            db_file_sequential_read = re.compile('db file sequential read')
            control_file_sequential_read = re.compile('control file sequential read')
            foregroud_waittime = {}

            for rawdata in soup.find_all('table'):
                if re.compile('This table displays foreground wait event information').search(rawdata['summary']):
                    trdatas = rawdata.find_all(name='tr')
                    for trdata in trdatas:
                        tddatas = BeautifulSoup(str(trdata), "html.parser")
                        tddata = tddatas.find_all(name='td')
                        try:
                            if (db_file_parallel_read.match(tddata[1].string)):
                                foregroud_waittime["db file parallel read Avg Wait time"] = tddata[5].string
                                foregroud_waittime["db file parallel read Waits"] = tddata[2].string
                            if (db_file_sequential_read.match(tddata[1].string)):
                                foregroud_waittime["db file sequential read Avg Wait time"] = tddata[5].string
                                foregroud_waittime["db file sequential read Waits"] = tddata[2].string
                            if (control_file_sequential_read.match(tddata[1].string)):
                                foregroud_waittime["control file sequential read Avg Wait time"] = tddata[5].string
                                foregroud_waittime["control file sequential read write Waits"] = tddata[2].string
                        except Exception, e:
                            pass
            # print foregroud_waittime
            f.close()

            # Backgroud Wait Time Metrics
            with open(awrfile) as f:
                soup = BeautifulSoup(f, 'lxml')

                # IO latency calculation
                log_file_parallel_write = re.compile('log file parallel write')
                control_file_parallel_write = re.compile('control file parallel write')
                db_file_paralle_write = re.compile('db file parallel write')
                backgroud_waittime = {}

                for rawdata in soup.find_all('table'):
                    if re.compile('This table displays background wait event information').search(rawdata['summary']):
                        trdatas = rawdata.find_all(name='tr')
                        for trdata in trdatas:
                            tddatas = BeautifulSoup(str(trdata), "html.parser")
                            tddata = tddatas.find_all(name='td')
                            try:
                                if (log_file_parallel_write.match(tddata[1].string)):
                                    backgroud_waittime["log file parallel write Avg Wait time"] = tddata[5].string
                                    backgroud_waittime["log file parallel write Waits"] = tddata[2].string
                                if (control_file_parallel_write.match(tddata[1].string)):
                                    backgroud_waittime["control file parallel write Avg Wait time"] = tddata[5].string
                                    backgroud_waittime["control file parallel write write Waits"] = tddata[2].string
                                if (db_file_paralle_write.match(tddata[1].string)):
                                    backgroud_waittime["db file parallel write Avg Wait time"] = tddata[5].string
                                    backgroud_waittime["db file parallel write write Waits"] = tddata[2].string
                            except Exception, e:
                                pass
                # print backgroud_waittime
                f.close()
                print dbname[0]["DB Name"]+" AWR Report parsing is finished"

    try:
        wb = load_workbook("awr_metrics.xlsx")
        ws = wb.get_sheet_by_name("awr_metric")

        ws.cell(row=line_no, column=3).value = dbname[0]["DB Name"]
        ws.cell(row=line_no, column=4).value = begin[0]["Begin SnapShot"] + "--" + endy[0]["End Snapshot"]
        ws.cell(row=line_no, column=6).value = dbcachesize
        for i in range(len(cpu_data) - 1):
            for key, value in cpu_data[i].items():
                ws.cell(row=line_no, column=7).value = cpu_data[1]["%Busy"]
                ws.cell(row=line_no, column=8).value = cpu_data[5]["%Busy"]
        ws.cell(row=line_no, column=9).value = float(result["physical read total IO requests"])
        ws.cell(row=line_no, column=10).value = float(result["physical write total IO requests"])
        ws.cell(row=line_no, column=11).value = float(result["physical read total IO requests"]) + float(
            result["physical write total IO requests"])
        ws.cell(row=line_no, column=12).value = foregroud_waittime["db file sequential read Waits"]
        ws.cell(row=line_no, column=13).value = foregroud_waittime["db file sequential read Avg Wait time"]
        ws.cell(row=line_no, column=14).value = foregroud_waittime["db file parallel read Waits"]
        ws.cell(row=line_no, column=15).value = foregroud_waittime["db file parallel read Avg Wait time"]
        ws.cell(row=line_no, column=16).value = backgroud_waittime["log file parallel write Waits"]
        ws.cell(row=line_no, column=17).value = backgroud_waittime["log file parallel write Avg Wait time"]
        ws.cell(row=line_no, column=18).value = backgroud_waittime["db file parallel write write Waits"]
        ws.cell(row=line_no, column=19).value = backgroud_waittime["db file parallel write Avg Wait time"]
        wb.save("awr_metrics.xlsx")
    except Exception, e:
        pass

        # csvfile=file(dbname[0]["DB Name"]+".csv",'ab+')
        # csvfile.truncate()
        # try:
        #     writer = csv.writer(csvfile,dialect='excel')
        #
        #
        #     writer.writerow(['DB Name', dbname[0]["DB Name"]])
        #     writer.writerow('\n')
        #     #,['End SnapShot', end[0]["End Snapshot"]]
        #     writer.writerow(['Begin SnapShot', begin[0]["Begin SnapShot"]])
        #     writer.writerow(['End SnapShot', endy[0]["End Snapshot"]])
        #     writer.writerow('\n')
        #
        #     writer.writerow(['IO Metric'])
        #     writer.writerow(['physical write total IO requests','physical read total IO requests'])
        #     writer.writerow([float(result["physical write total IO requests"]), float(result["physical read total IO requests"])])
        #
        #     writer.writerow(['physical write total bytes', 'physical read total bytes'])
        #     writer.writerow([float(result["physical write total bytes"]), float(result["physical read total bytes"])])
        #
        #     writer.writerow('\n')
        #
        #     writer.writerow(['IO latency'])
        #     writer.writerow(['IO Wait Event','Avg Wait Time','Waits'])
        #     writer.writerow(['db file parallel read',foregroud_waittime["db file parallel read Avg Wait time"], foregroud_waittime["db file parallel read Waits"]])
        #     writer.writerow(['db file sequential read', foregroud_waittime["db file sequential read Avg Wait time"],
        #                      foregroud_waittime["db file sequential read Waits"]])
        #     writer.writerow(['control file sequential read', foregroud_waittime["control file sequential read Avg Wait time"],
        #                      foregroud_waittime["control file sequential read write Waits"]])
        #     writer.writerow(['log file parallel write', backgroud_waittime["log file parallel write Avg Wait time"],
        #                      backgroud_waittime["log file parallel write Waits"]])
        #     writer.writerow(['control file parallel write', backgroud_waittime["control file parallel write Avg Wait time"],
        #                      backgroud_waittime["control file parallel write write Waits"]])
        #
        #     writer.writerow('\n')
        #     writer.writerow(['CPU utilization'])
        #
        #     for i in range(len(cpu_data)-1):
        #         for key, value in cpu_data[i].items():
        #             writer.writerow([key,value])
        # except Exception, e:
        #     pass
        # csvfile.close()


def extrace_from_directory(awrdir):

    files = glob.iglob(awrdir+'\*.html')

    # The actual data in Excel Template at the third row
    line_no = 3
    for item in files:
        extrace_from_file(item,line_no)
        line_no +=1

def clear_data():
    try:
        wb = load_workbook("awr_metrics.xlsx")
        ws = wb.get_sheet_by_name("awr_metric")
        # clear excel data
        maxrow = ws.max_row
        for i in range(3,maxrow+1):
            ws.cell(row=i, column=3).value = ''
            ws.cell(row=i, column=4).value = ''
            ws.cell(row=i, column=5).value = ''
            ws.cell(row=i, column=6).value = ''
            ws.cell(row=i, column=7).value = ''
            ws.cell(row=i, column=8).value = ''
            ws.cell(row=i, column=9).value = ''
            ws.cell(row=i, column=10).value = ''
            ws.cell(row=i, column=11).value = ''
            ws.cell(row=i, column=12).value = ''
            ws.cell(row=i, column=13).value = ''
            ws.cell(row=i, column=14).value = ''
            ws.cell(row=i, column=15).value = ''
            ws.cell(row=i, column=16).value = ''
            ws.cell(row=i, column=17).value = ''
            ws.cell(row=i, column=18).value = ''
            ws.cell(row=i, column=19).value = ''
        wb.save("awr_metrics.xlsx")
    except Exception, e:
        print e

def extract_data(**kwargs):

    if kwargs["file"] is not None:
        awrfile = os.path.join(DATADIR, kwargs["file"])
        extrace_from_file(awrfile,3)
    if  kwargs["directory"] is not None:
        awrdir = kwargs["directory"]
        extrace_from_directory(awrdir)


def _argparser():
    parser = argparse.ArgumentParser(description='AWR Data Analyze')
    parser.add_argument('--f','--file', action='store', dest='file', help='AWR File Name')
    parser.add_argument('--d', '--directory', action='store', dest='directory', help='AWR File Directory')

    return parser.parse_args()

if __name__ == "__main__":
    parser = _argparser()
    if parser.file is None and parser.directory is None:
        print "usage: fetchmetric_awr.py [-h] [--f FILE] [--d DIRECTORY]"
    conn_args = dict(file=parser.file,directory=parser.directory)
    clear_data()
    extract_data(**conn_args)

